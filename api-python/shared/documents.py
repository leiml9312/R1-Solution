# Builders for the three formatted business documents (Invoice, Packing
# List, Statement) modeled on the company's existing Excel template —
# see the header image (company-header.png) embedded in every export.
#
# Each build_*_xlsx() returns an in-memory .xlsx (openpyxl, with real
# formulas so totals recalculate if the file is edited later) and each
# build_*_pdf() returns an in-memory .pdf (reportlab) rendering of the
# same data for cases where a static, ready-to-send file is wanted.

import io
import os
import re
from datetime import date, datetime
from xml.sax.saxutils import escape as xml_escape

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, Side

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import Image as RLImage
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

_THIS_DIR = os.path.dirname(os.path.abspath(__file__))
HEADER_IMAGE_PATH = os.path.join(_THIS_DIR, "company-header.png")
HEADER_IMAGE_ASPECT = 200 / 1286  # native company-header.png is 1286x200px

# --- shared styling -----------------------------------------------------

FONT_TITLE = Font(name="Arial", size=20, bold=True)
FONT_LABEL = Font(name="Arial", size=11, bold=True)
FONT_HEADER = Font(name="Arial", size=10, bold=True)
FONT_BODY = Font(name="Arial", size=10)
FONT_BOLD_BODY = Font(name="Arial", size=10, bold=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)
_THIN = Side(style="thin")
BORDER_ALL = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_STYLES = getSampleStyleSheet()
STYLE_NORMAL = ParagraphStyle("body", parent=_STYLES["Normal"], fontName="Helvetica", fontSize=9, leading=12)
STYLE_BOLD = ParagraphStyle("bold", parent=_STYLES["Normal"], fontName="Helvetica-Bold", fontSize=9, leading=12)
STYLE_TITLE = ParagraphStyle(
    "title", parent=_STYLES["Heading1"], fontName="Helvetica-Bold", fontSize=20, alignment=1
)
STYLE_HEADER_CELL = ParagraphStyle(
    "headerCell", parent=_STYLES["Normal"], fontName="Helvetica-Bold", fontSize=8, alignment=1
)


def _num(value, default=0.0):
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def esc(value) -> str:
    return xml_escape(str(value if value is not None else ""))


def safe_filename(value, fallback):
    value = (value or "").strip()
    if not value:
        return fallback
    return re.sub(r"[^A-Za-z0-9._-]+", "_", value)


def _parse_date(value):
    if not value:
        return None
    if isinstance(value, (datetime, date)):
        return value
    for fmt in ("%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(str(value), fmt).date()
        except ValueError:
            continue
    return value


def _fmt_qty(value) -> str:
    return f"{value:g}"


# --- header image (letterhead) ------------------------------------------


def add_xlsx_header(ws, header_rows=4, row_height=22, img_width=700):
    """Anchors the company letterhead at A1 and reserves rows for it.
    Returns the first free row below the header."""
    img = XLImage(HEADER_IMAGE_PATH)
    img.width = img_width
    img.height = int(img_width * HEADER_IMAGE_ASPECT)
    ws.add_image(img, "A1")
    for r in range(1, header_rows + 1):
        ws.row_dimensions[r].height = row_height
    return header_rows + 1


def _pdf_header_flowables(usable_width):
    img_height = usable_width * HEADER_IMAGE_ASPECT
    return [RLImage(HEADER_IMAGE_PATH, width=usable_width, height=img_height), Spacer(1, 12)]


def _new_pdf_doc(buffer):
    return SimpleDocTemplate(buffer, pagesize=letter, topMargin=36, bottomMargin=36, leftMargin=36, rightMargin=36)


# =========================================================================
# Invoice
# =========================================================================


def build_invoice_xlsx(data: dict) -> io.BytesIO:
    data = data or {}
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice"
    ws.sheet_view.showGridLines = False
    for col, width in zip("ABCDEFGH", [15, 6, 10, 36, 9, 12, 13, 14]):
        ws.column_dimensions[col].width = width

    row = add_xlsx_header(ws)
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    title = ws.cell(row=row, column=1, value="INVOICE")
    title.font = FONT_TITLE
    title.alignment = ALIGN_CENTER
    row += 2

    ship_row = row
    ws.merge_cells(start_row=ship_row, start_column=1, end_row=ship_row, end_column=4)
    ws.merge_cells(start_row=ship_row, start_column=5, end_row=ship_row, end_column=8)
    ship_cell = ws.cell(row=ship_row, column=1, value=f"Ship To: {data.get('shipTo', '')}")
    ship_cell.font = FONT_BODY
    ship_cell.alignment = ALIGN_LEFT
    date_cell = ws.cell(row=ship_row, column=5, value=f"DATE: {data.get('date', '')}")
    date_cell.font = FONT_BODY
    date_cell.alignment = ALIGN_RIGHT
    ws.row_dimensions[ship_row].height = 40
    row += 1

    bill_row = row
    ws.merge_cells(start_row=bill_row, start_column=1, end_row=bill_row, end_column=4)
    ws.merge_cells(start_row=bill_row, start_column=5, end_row=bill_row, end_column=8)
    bill_cell = ws.cell(row=bill_row, column=1, value=f"Bill To: {data.get('billTo', '')}")
    bill_cell.font = FONT_BODY
    bill_cell.alignment = ALIGN_LEFT
    inv_cell = ws.cell(row=bill_row, column=5, value=f"INVOICE NO: {data.get('invoiceNo', '')}")
    inv_cell.font = FONT_LABEL
    inv_cell.alignment = ALIGN_RIGHT
    ws.row_dimensions[bill_row].height = 40
    row += 2

    headers = ["PO NO.", "Line", "PART NO.", "DESCRIPTION", "QTY\nPCS", "UNIT PRICE\nUSD", "AMOUNT\nUSD", "D/C"]
    header_row = row
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=i, value=h)
        c.font = FONT_HEADER
        c.alignment = ALIGN_CENTER
        c.border = BORDER_ALL
    ws.row_dimensions[header_row].height = 30
    row += 1

    line_items = data.get("lineItems") or []
    first_item_row = row
    for idx, item in enumerate(line_items, start=1):
        values = [
            item.get("poNo", ""),
            item.get("line", idx),
            item.get("partNo", ""),
            item.get("description", ""),
            _num(item.get("qty")),
            _num(item.get("unitPrice")),
            None,
            item.get("dc", ""),
        ]
        for col, v in enumerate(values, start=1):
            c = ws.cell(row=row, column=col, value=v)
            c.font = FONT_BODY
            c.alignment = ALIGN_LEFT if col == 4 else ALIGN_CENTER
            c.border = BORDER_ALL
        ws.cell(row=row, column=6).number_format = "0.00"
        amount_cell = ws.cell(row=row, column=7, value=f"=E{row}*F{row}")
        amount_cell.number_format = "#,##0.00"
        row += 1
    if not line_items:
        for col in range(1, 9):
            ws.cell(row=row, column=col).border = BORDER_ALL
        row += 1
    last_item_row = row - 1

    total_row = row
    label_cell = ws.cell(row=total_row, column=4, value="Total")
    label_cell.font = FONT_BOLD_BODY
    label_cell.alignment = ALIGN_RIGHT
    qty_total = ws.cell(row=total_row, column=5, value=f"=SUM(E{first_item_row}:E{last_item_row})")
    qty_total.font = FONT_BOLD_BODY
    qty_total.alignment = ALIGN_CENTER
    amt_total = ws.cell(row=total_row, column=7, value=f"=SUM(G{first_item_row}:G{last_item_row})")
    amt_total.font = FONT_BOLD_BODY
    amt_total.alignment = ALIGN_CENTER
    amt_total.number_format = "#,##0.00"
    for col in range(1, 9):
        ws.cell(row=total_row, column=col).border = BORDER_ALL
    row += 1

    terms = data.get("terms") or {}
    ws.cell(row=row, column=1, value=terms.get("incoterm", "EXW SZ")).font = FONT_BODY
    ws.cell(row=row, column=3, value=terms.get("shipMethod", "By Sea")).font = FONT_BODY
    ws.cell(row=row, column=4, value=f"HS Code: {terms.get('hsCode', '')}").font = FONT_BODY
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws.cell(row=row, column=1, value=f"Payment Term: {data.get('paymentTerm', '')}").font = FONT_BODY
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws.cell(row=row, column=1, value=f"Remark: {data.get('remark', '')}").font = FONT_BODY

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def build_invoice_pdf(data: dict) -> io.BytesIO:
    data = data or {}
    buffer = io.BytesIO()
    doc = _new_pdf_doc(buffer)
    usable_width = doc.width
    elements = _pdf_header_flowables(usable_width)
    elements.append(Paragraph("INVOICE", STYLE_TITLE))
    elements.append(Spacer(1, 10))

    info = Table(
        [
            [
                Paragraph(f"<b>Ship To:</b> {esc(data.get('shipTo', ''))}", STYLE_NORMAL),
                Paragraph(f"<b>DATE:</b> {esc(data.get('date', ''))}", STYLE_NORMAL),
            ],
            [
                Paragraph(f"<b>Bill To:</b> {esc(data.get('billTo', ''))}", STYLE_NORMAL),
                Paragraph(f"<b>INVOICE NO:</b> {esc(data.get('invoiceNo', ''))}", STYLE_NORMAL),
            ],
        ],
        colWidths=[usable_width * 0.65, usable_width * 0.35],
    )
    info.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP"), ("BOTTOMPADDING", (0, 0), (-1, -1), 6)]))
    elements.append(info)
    elements.append(Spacer(1, 10))

    headers = ["PO NO.", "Line", "PART NO.", "DESCRIPTION", "QTY\nPCS", "UNIT PRICE\nUSD", "AMOUNT\nUSD", "D/C"]
    rows = [[Paragraph(h.replace("\n", "<br/>"), STYLE_HEADER_CELL) for h in headers]]
    qty_total = 0.0
    amt_total = 0.0
    for idx, item in enumerate(data.get("lineItems") or [], start=1):
        qty = _num(item.get("qty"))
        price = _num(item.get("unitPrice"))
        amount = qty * price
        qty_total += qty
        amt_total += amount
        rows.append(
            [
                item.get("poNo", ""),
                str(item.get("line", idx)),
                str(item.get("partNo", "")),
                Paragraph(esc(item.get("description", "")), STYLE_NORMAL),
                _fmt_qty(qty),
                f"{price:,.2f}",
                f"{amount:,.2f}",
                item.get("dc", ""),
            ]
        )
    rows.append(["", "", "", "Total", _fmt_qty(qty_total), "", f"{amt_total:,.2f}", ""])

    col_widths = [usable_width * w for w in [0.12, 0.05, 0.10, 0.30, 0.09, 0.12, 0.12, 0.10]]
    table = Table(rows, colWidths=col_widths, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                ("BACKGROUND", (0, 0), (-1, 0), colors.whitesmoke),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("FONTNAME", (3, -1), (3, -1), "Helvetica-Bold"),
            ]
        )
    )
    elements.append(table)
    elements.append(Spacer(1, 14))

    terms = data.get("terms") or {}
    elements.append(
        Paragraph(
            f"{esc(terms.get('incoterm', 'EXW SZ'))} &nbsp;&nbsp; {esc(terms.get('shipMethod', 'By Sea'))} "
            f"&nbsp;&nbsp; HS Code: {esc(terms.get('hsCode', ''))}",
            STYLE_NORMAL,
        )
    )
    elements.append(Spacer(1, 6))
    elements.append(Paragraph(f"<b>Payment Term:</b> {esc(data.get('paymentTerm', ''))}", STYLE_NORMAL))
    elements.append(Spacer(1, 6))
    elements.append(Paragraph(f"<b>Remark:</b> {esc(data.get('remark', ''))}", STYLE_NORMAL))

    doc.build(elements)
    buffer.seek(0)
    return buffer


# =========================================================================
# Packing List
# =========================================================================


def build_packing_list_xlsx(data: dict) -> io.BytesIO:
    data = data or {}
    wb = Workbook()
    ws = wb.active
    ws.title = "Packing List"
    ws.sheet_view.showGridLines = False
    for col, width in zip("ABCDEFGH", [10, 10, 10, 34, 10, 10, 12, 14]):
        ws.column_dimensions[col].width = width

    row = add_xlsx_header(ws)
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    title = ws.cell(row=row, column=1, value="PACKING LIST")
    title.font = FONT_TITLE
    title.alignment = ALIGN_CENTER
    row += 2

    ship_row = row
    ws.merge_cells(start_row=ship_row, start_column=1, end_row=ship_row, end_column=5)
    ws.merge_cells(start_row=ship_row, start_column=6, end_row=ship_row, end_column=8)
    ws.cell(row=ship_row, column=1, value=f"Ship To: {data.get('shipTo', '')}").font = FONT_BODY
    ws.cell(row=ship_row, column=1).alignment = ALIGN_LEFT
    no_cell = ws.cell(row=ship_row, column=6, value=f"NO.: {data.get('no', '')}")
    no_cell.font = FONT_BODY
    no_cell.alignment = ALIGN_RIGHT
    ws.row_dimensions[ship_row].height = 40
    row += 1

    bill_row = row
    ws.merge_cells(start_row=bill_row, start_column=1, end_row=bill_row, end_column=5)
    ws.merge_cells(start_row=bill_row, start_column=6, end_row=bill_row, end_column=8)
    ws.cell(row=bill_row, column=1, value=f"Bill To: {data.get('billTo', '')}").font = FONT_BODY
    ws.cell(row=bill_row, column=1).alignment = ALIGN_LEFT
    date_cell = ws.cell(row=bill_row, column=6, value=f"DATE: {data.get('date', '')}")
    date_cell.font = FONT_BODY
    date_cell.alignment = ALIGN_RIGHT
    ws.row_dimensions[bill_row].height = 30
    row += 2

    header_row1, header_row2 = row, row + 1
    row1_labels = {
        1: "Packing No",
        2: "Packing No",
        3: "PART NO.",
        4: "DESCRIPTION",
        5: "Quantity",
        6: "Net weight",
        7: "Gross weight",
        8: "Measurement",
    }
    row2_labels = {1: "Pallet NO.", 2: "CTN NO.", 5: "PIECES", 6: "KGS", 7: "KGS", 8: "CM"}
    for col, label in row1_labels.items():
        c = ws.cell(row=header_row1, column=col, value=label)
        c.font = FONT_HEADER
        c.alignment = ALIGN_CENTER
        c.border = BORDER_ALL
    for col in (3, 4):
        ws.merge_cells(start_row=header_row1, start_column=col, end_row=header_row2, end_column=col)
    for col in row1_labels:
        c2 = ws.cell(row=header_row2, column=col, value=row2_labels.get(col))
        c2.font = FONT_HEADER
        c2.alignment = ALIGN_CENTER
        c2.border = BORDER_ALL
    row = header_row2 + 1

    line_items = data.get("lineItems") or []
    first_item_row = row
    pallet_groups = []
    current_pallet, group_start = None, None
    for idx, item in enumerate(line_items):
        r = row + idx
        pallet_no = item.get("palletNo", "")
        values = [
            pallet_no,
            _num(item.get("ctnNo")),
            item.get("partNo", ""),
            item.get("description", ""),
            _num(item.get("quantity")),
            _num(item.get("netWeight")),
            _num(item.get("grossWeight")),
            item.get("measurement", ""),
        ]
        for col, v in enumerate(values, start=1):
            c = ws.cell(row=r, column=col, value=v)
            c.font = FONT_BODY
            c.alignment = ALIGN_LEFT if col == 4 else ALIGN_CENTER
            c.border = BORDER_ALL
        if pallet_no != current_pallet:
            if current_pallet is not None:
                pallet_groups.append((group_start, r - 1))
            current_pallet = pallet_no
            group_start = r
    if line_items:
        pallet_groups.append((group_start, row + len(line_items) - 1))
        last_item_row = row + len(line_items) - 1
    else:
        last_item_row = row
        for col in range(1, 9):
            ws.cell(row=row, column=col).border = BORDER_ALL

    for start_r, end_r in pallet_groups:
        if end_r > start_r:
            ws.merge_cells(start_row=start_r, start_column=1, end_row=end_r, end_column=1)
            measurements = {ws.cell(row=r, column=8).value for r in range(start_r, end_r + 1)}
            if len(measurements) == 1:
                ws.merge_cells(start_row=start_r, start_column=8, end_row=end_r, end_column=8)
    row = last_item_row + 1

    total_row = row
    ws.cell(row=total_row, column=2, value=f"=SUM(B{first_item_row}:B{last_item_row})")
    label_cell = ws.cell(row=total_row, column=4, value="Total")
    label_cell.alignment = ALIGN_RIGHT
    ws.cell(row=total_row, column=5, value=f"=SUM(E{first_item_row}:E{last_item_row})")
    ws.cell(row=total_row, column=6, value=f"=SUM(F{first_item_row}:F{last_item_row})")
    ws.cell(row=total_row, column=7, value=f"=SUM(G{first_item_row}:G{last_item_row})")
    for col in range(1, 9):
        c = ws.cell(row=total_row, column=col)
        c.font = FONT_BOLD_BODY
        c.border = BORDER_ALL
        if col not in (4,):
            c.alignment = ALIGN_CENTER
    for col in (6, 7):
        ws.cell(row=total_row, column=col).number_format = "#,##0.0"
    row = total_row + 1

    distinct_pallets = len({item.get("palletNo", "") for item in line_items}) if line_items else 0
    total_ctns = sum(_num(item.get("ctnNo")) for item in line_items)
    total_ctns_display = int(total_ctns) if total_ctns == int(total_ctns) else total_ctns
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    summary_cell = ws.cell(row=row, column=1, value=f"TOTAL: {distinct_pallets} PLT({total_ctns_display} CTNS)")
    summary_cell.font = FONT_LABEL

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def build_packing_list_pdf(data: dict) -> io.BytesIO:
    data = data or {}
    buffer = io.BytesIO()
    doc = _new_pdf_doc(buffer)
    usable_width = doc.width
    elements = _pdf_header_flowables(usable_width)
    elements.append(Paragraph("PACKING LIST", STYLE_TITLE))
    elements.append(Spacer(1, 10))

    info = Table(
        [
            [
                Paragraph(f"<b>Ship To:</b> {esc(data.get('shipTo', ''))}", STYLE_NORMAL),
                Paragraph(f"<b>NO.:</b> {esc(data.get('no', ''))}", STYLE_NORMAL),
            ],
            [
                Paragraph(f"<b>Bill To:</b> {esc(data.get('billTo', ''))}", STYLE_NORMAL),
                Paragraph(f"<b>DATE:</b> {esc(data.get('date', ''))}", STYLE_NORMAL),
            ],
        ],
        colWidths=[usable_width * 0.65, usable_width * 0.35],
    )
    info.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP"), ("BOTTOMPADDING", (0, 0), (-1, -1), 6)]))
    elements.append(info)
    elements.append(Spacer(1, 10))

    headers = [
        "Pallet\nNO.",
        "CTN\nNO.",
        "PART NO.",
        "DESCRIPTION",
        "Quantity\n(PCS)",
        "Net Wt.\n(KGS)",
        "Gross Wt.\n(KGS)",
        "Measurement\n(CM)",
    ]
    rows = [[Paragraph(h.replace("\n", "<br/>"), STYLE_HEADER_CELL) for h in headers]]
    ctn_total = qty_total = net_total = gross_total = 0.0
    pallets = set()
    for item in data.get("lineItems") or []:
        ctn = _num(item.get("ctnNo"))
        qty = _num(item.get("quantity"))
        net = _num(item.get("netWeight"))
        gross = _num(item.get("grossWeight"))
        ctn_total += ctn
        qty_total += qty
        net_total += net
        gross_total += gross
        pallets.add(item.get("palletNo", ""))
        rows.append(
            [
                item.get("palletNo", ""),
                _fmt_qty(ctn),
                item.get("partNo", ""),
                Paragraph(esc(item.get("description", "")), STYLE_NORMAL),
                _fmt_qty(qty),
                f"{net:,.1f}",
                f"{gross:,.1f}",
                item.get("measurement", ""),
            ]
        )
    rows.append(
        ["", _fmt_qty(ctn_total), "", "Total", _fmt_qty(qty_total), f"{net_total:,.1f}", f"{gross_total:,.1f}", ""]
    )

    col_widths = [usable_width * w for w in [0.08, 0.08, 0.11, 0.30, 0.11, 0.11, 0.11, 0.10]]
    table = Table(rows, colWidths=col_widths, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                ("BACKGROUND", (0, 0), (-1, 0), colors.whitesmoke),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("FONTNAME", (3, -1), (3, -1), "Helvetica-Bold"),
            ]
        )
    )
    elements.append(table)
    elements.append(Spacer(1, 12))

    total_ctns_display = int(ctn_total) if ctn_total == int(ctn_total) else ctn_total
    summary = f"TOTAL: {len(pallets)} PLT({total_ctns_display} CTNS)"
    elements.append(Paragraph(f"<b>{esc(summary)}</b>", STYLE_NORMAL))

    doc.build(elements)
    buffer.seek(0)
    return buffer


# =========================================================================
# Statement
# =========================================================================


def build_statement_xlsx(data: dict) -> io.BytesIO:
    data = data or {}
    wb = Workbook()
    ws = wb.active
    ws.title = "Statement"
    ws.sheet_view.showGridLines = False
    for col, width in zip("ABCDE", [14, 18, 10, 12, 32]):
        ws.column_dimensions[col].width = width

    row = add_xlsx_header(ws)
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    title = ws.cell(row=row, column=1, value="STATEMENT")
    title.font = FONT_TITLE
    title.alignment = ALIGN_CENTER
    row += 2

    to_row = row
    ws.cell(row=to_row, column=1, value="TO:").font = FONT_LABEL
    to_cell = ws.cell(row=to_row, column=2, value=data.get("to", ""))
    to_cell.font = FONT_BODY
    to_cell.alignment = ALIGN_LEFT
    ws.cell(row=to_row, column=4, value="DATE:").font = FONT_LABEL
    date_cell = ws.cell(row=to_row, column=5, value=_parse_date(data.get("date")))
    date_cell.number_format = "yyyy/mm/dd"
    date_cell.font = FONT_BODY
    ws.row_dimensions[to_row].height = 40
    row += 1

    if data.get("attn"):
        ws.cell(row=row, column=1, value="ATTN:").font = FONT_LABEL
        ws.cell(row=row, column=2, value=data["attn"]).font = FONT_BODY
        row += 1
    if data.get("tel"):
        ws.cell(row=row, column=1, value="TEL:").font = FONT_LABEL
        ws.cell(row=row, column=2, value=data["tel"]).font = FONT_BODY
        row += 1

    ws.cell(row=row, column=4, value="From:").font = FONT_LABEL
    ws.cell(row=row, column=5, value=data.get("from", "")).font = FONT_BODY
    row += 2

    headers = ["Date", "Invoice No.", "Currency", "Amount", "Remark"]
    header_row = row
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=i, value=h)
        c.font = FONT_HEADER
        c.alignment = ALIGN_CENTER
        c.border = BORDER_ALL
    row += 1

    line_items = data.get("lineItems") or []
    first_item_row = row
    for item in line_items:
        date_cell = ws.cell(row=row, column=1, value=_parse_date(item.get("date")))
        date_cell.number_format = "yyyy/mm/dd"
        ws.cell(row=row, column=2, value=item.get("invoiceNo", ""))
        ws.cell(row=row, column=3, value=item.get("currency", "USD"))
        amount_cell = ws.cell(row=row, column=4, value=_num(item.get("amount")))
        amount_cell.number_format = "#,##0.00"
        ws.cell(row=row, column=5, value=item.get("remark", ""))
        for col in range(1, 6):
            c = ws.cell(row=row, column=col)
            c.font = FONT_BODY
            c.alignment = ALIGN_CENTER
            c.border = BORDER_ALL
        row += 1
    if not line_items:
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = BORDER_ALL
        row += 1
    last_item_row = row - 1

    total_row = row
    total_label = ws.cell(row=total_row, column=3, value="TOTAL")
    total_label.font = FONT_BOLD_BODY
    total_label.alignment = ALIGN_RIGHT
    total_cell = ws.cell(row=total_row, column=4, value=f"=SUM(D{first_item_row}:D{last_item_row})")
    total_cell.font = FONT_BOLD_BODY
    total_cell.number_format = "#,##0.00"
    total_cell.alignment = ALIGN_CENTER
    for col in range(1, 6):
        ws.cell(row=total_row, column=col).border = BORDER_ALL
    row += 1

    aging_header_row = row
    for i, label in enumerate(["NOT YET  DUE", "OVER 30 DAYS", "OVER 60 DAYS", "OUTSTANDING TOTAL"], start=1):
        c = ws.cell(row=aging_header_row, column=i, value=label)
        c.font = FONT_HEADER
        c.alignment = ALIGN_CENTER
        c.border = BORDER_ALL
    row += 1

    aging_row = row
    date_range = f"A{first_item_row}:A{last_item_row}"
    amt_range = f"D{first_item_row}:D{last_item_row}"
    ref = f"$E${to_row}"
    formulas = [
        f'=SUMIFS({amt_range},{date_range},">="&({ref}-30))',
        f'=SUMIFS({amt_range},{date_range},"<"&({ref}-30),{date_range},">="&({ref}-60))',
        f'=SUMIFS({amt_range},{date_range},"<"&({ref}-60))',
        None,  # outstanding total set below, referencing the three cells just written
    ]
    for col, formula in enumerate(formulas[:3], start=1):
        c = ws.cell(row=aging_row, column=col, value=formula)
        c.number_format = "#,##0.00"
        c.font = FONT_BODY
        c.alignment = ALIGN_CENTER
        c.border = BORDER_ALL
    outstanding = ws.cell(row=aging_row, column=4, value=f"=A{aging_row}+B{aging_row}+C{aging_row}")
    outstanding.number_format = "#,##0.00"
    outstanding.font = FONT_BODY
    outstanding.alignment = ALIGN_CENTER
    outstanding.border = BORDER_ALL
    row += 1

    if data.get("paymentTerm"):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws.cell(row=row, column=1, value=f"Payment Term: {data['paymentTerm']}").font = FONT_BODY
        row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    remark = data.get("remark") or (
        "Kindly please confirm the statement within 7 working days, otherwise it will be regard as acceptance."
    )
    ws.cell(row=row, column=1, value=f"Remark: {remark}").font = FONT_BODY

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def build_statement_pdf(data: dict) -> io.BytesIO:
    data = data or {}
    buffer = io.BytesIO()
    doc = _new_pdf_doc(buffer)
    usable_width = doc.width
    elements = _pdf_header_flowables(usable_width)
    elements.append(Paragraph("STATEMENT", STYLE_TITLE))
    elements.append(Spacer(1, 10))

    attn_line = f"<b>ATTN:</b> {esc(data.get('attn', ''))}" if data.get("attn") else ""
    info = Table(
        [
            [
                Paragraph(f"<b>TO:</b> {esc(data.get('to', ''))}", STYLE_NORMAL),
                Paragraph(f"<b>DATE:</b> {esc(data.get('date', ''))}", STYLE_NORMAL),
            ],
            [
                Paragraph(attn_line, STYLE_NORMAL),
                Paragraph(f"<b>From:</b> {esc(data.get('from', ''))}", STYLE_NORMAL),
            ],
        ],
        colWidths=[usable_width * 0.65, usable_width * 0.35],
    )
    info.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP"), ("BOTTOMPADDING", (0, 0), (-1, -1), 6)]))
    elements.append(info)
    elements.append(Spacer(1, 10))

    headers = ["Date", "Invoice No.", "Currency", "Amount", "Remark"]
    rows = [[Paragraph(h, STYLE_HEADER_CELL) for h in headers]]
    total = 0.0
    aging = {"notYetDue": 0.0, "over30": 0.0, "over60": 0.0}
    ref_date = _parse_date(data.get("date"))
    ref_date = ref_date if isinstance(ref_date, date) else date.today()
    for item in data.get("lineItems") or []:
        amount = _num(item.get("amount"))
        total += amount
        item_date = _parse_date(item.get("date"))
        age = (ref_date - item_date).days if isinstance(item_date, date) else 0
        if age <= 30:
            aging["notYetDue"] += amount
        elif age <= 60:
            aging["over30"] += amount
        else:
            aging["over60"] += amount
        rows.append(
            [
                item.get("date", ""),
                item.get("invoiceNo", ""),
                item.get("currency", "USD"),
                f"{amount:,.2f}",
                Paragraph(esc(item.get("remark", "")), STYLE_NORMAL),
            ]
        )
    rows.append(["", "", "TOTAL", f"{total:,.2f}", ""])

    col_widths = [usable_width * w for w in [0.16, 0.24, 0.14, 0.16, 0.30]]
    table = Table(rows, colWidths=col_widths, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                ("BACKGROUND", (0, 0), (-1, 0), colors.whitesmoke),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("FONTNAME", (2, -1), (3, -1), "Helvetica-Bold"),
            ]
        )
    )
    elements.append(table)
    elements.append(Spacer(1, 14))

    aging_rows = [
        ["NOT YET DUE", "OVER 30 DAYS", "OVER 60 DAYS", "OUTSTANDING TOTAL"],
        [f"{aging['notYetDue']:,.2f}", f"{aging['over30']:,.2f}", f"{aging['over60']:,.2f}", f"{total:,.2f}"],
    ]
    aging_table = Table(aging_rows, colWidths=[usable_width * 0.25] * 4)
    aging_table.setStyle(
        TableStyle(
            [
                ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                ("BACKGROUND", (0, 0), (-1, 0), colors.whitesmoke),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ]
        )
    )
    elements.append(aging_table)
    elements.append(Spacer(1, 14))

    if data.get("paymentTerm"):
        elements.append(Paragraph(f"<b>Payment Term:</b> {esc(data['paymentTerm'])}", STYLE_NORMAL))
        elements.append(Spacer(1, 6))
    remark = data.get("remark") or (
        "Kindly please confirm the statement within 7 working days, otherwise it will be regard as acceptance."
    )
    elements.append(Paragraph(f"<b>Remark:</b> {esc(remark)}", STYLE_NORMAL))

    doc.build(elements)
    buffer.seek(0)
    return buffer
